#!/usr/bin/env python3
"""
Build Gastos_Final.xlsx — comprehensive Excel workbook with all historical expense data.
"""

import json
import copy
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = "/Users/guidocarminatti/Projects/playground/Gastos_Final.xlsx"

# ─── COLORS ────────────────────────────────────────────────────────────────────
C_NAVY      = "1A1A2E"
C_DARKBLUE  = "16213E"
C_ACCENT    = "0F3460"
C_GREEN     = "27AE60"
C_RED       = "E74C3C"
C_WHITE     = "FFFFFF"
C_YELLOW    = "F5A623"
C_LIGHTGRAY = "F4F5F7"
C_ALTROW    = "E8EAF0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=C_WHITE, bold=False, size=11, name="Calibri"):
    return Font(color=color, bold=bold, size=size, name=name)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── DATA ─────────────────────────────────────────────────────────────────────

OLA_DATA = {
    'Mar 2025': {
        'rate': 1000,
        'income': [('Basecoat', None, 933), ('Bono Compu Basecoat', None, 500), ('Basecoat 2', None, 963), ('Mati sun', None, 65), ('Fiesta Maru Bo', None, 65), ('Grego Dj', None, 25)],
        'fixed': [('MSI Cyborg', None, 810), ('Asisa Seguro', None, 45), ('Sillon', 230000, None), ('Monitor', 250000, None), ('AGIP', None, None), ('Exp Arenales', None, None), ('AYSA Arenales', None, None), ('Luz Edesur', None, None), ('Gas Arenales', None, None), ('Telecentro', None, None)],
        'daily': [],
    },
    'Apr 2025': {
        'rate': 1000,
        'income': [('Basecoat', None, 880), ('Basecoat 2', None, 963), ('Mati sun', None, 65), ('Fiesta Maru Bo', None, 65), ('Grego Dj', None, 25)],
        'fixed': [('Swiss', 150000, None), ('Telecentro', 36000, None), ('Fonoaudiologo', 20000, None), ('Gym', 23000, None), ('Asisa', None, 50), ('Santander Esp', None, 20), ('Olav Depto', 100000, None), ('Comida Olav', 500000, None)],
        'daily': [],
    },
    'May 2025': {
        'rate': 1000,
        'income': [('Basecoat', None, 880), ('Basecoat 2', None, 963)],
        'fixed': [('Swiss', 150000, None), ('Fonoaudiologo', 20000, None), ('Gym', 23000, None), ('Asisa', None, 50), ('Santander Esp', None, 20), ('Olav Depto', 100000, None), ('Comida Olav', 500000, None), ('Psico', 30000, None)],
        'daily': [('02', 'Bizcochos', 5300, None), ('04', 'Pietra milanesas', 13400, None), ('05', 'Psico mari', 20000, None), ('06', 'Psico 2', 20000, None), ('07', 'Psico 3 swiss', 13500, None)],
    },
    'Jun 2025': {
        'rate': 1100,
        'income': [('Basecoat', None, 880), ('Basecoat 2', None, 963)],
        'fixed': [('Swiss', 150000, None), ('Fonoaudiologo', 20000, None), ('Gym', 30000, None), ('Asisa', None, 50), ('Santander Esp', None, 20), ('Olav Depto', 100000, None), ('Comida Olav', 500000, None), ('Psico x4', 42000, None)],
        'daily': [('02', 'Bizcochos', 5300, None), ('03', 'Crazy cheese', 18000, None)],
    },
    'Jul 2025': {
        'rate': 1200,
        'income': [('Ipad Mini Bachi', None, 100)],
        'fixed': [('Swiss', 150000, None), ('Remedios', 70000, None), ('Gym', 30000, None), ('Asisa', None, 50), ('Santander Esp', None, 20), ('Olav Depto', 100000, None), ('Comida Olav', 500000, None), ('Psico x4', 42000, None)],
        'daily': [('02', 'Bizcochos', 5300, None), ('03', 'Crazy cheese', 18000, None), ('04', 'Auto Fabi', 30000, None), ('05', 'Empanadas', 45000, None), ('07', 'Strange', 30000, None)],
    },
    'Aug 2025': {
        'rate': 1300,
        'income': [],
        'fixed': [('Swiss', 150000, None), ('Remedios', 30000, None), ('Gym', 30000, None), ('Asisa', 50000, None), ('Santander Esp', 20000, None), ('Olav Depto', 130000, None), ('Comida Olav', 800000, None), ('Psico x2', 33000, None), ('Escribano', 30000, None), ('Internet', 20000, None)],
        'daily': [('02', 'Bizcochos', 5300, None), ('04', 'Auto Fabi', 30000, None), ('05', 'Empanadas', 45000, None)],
    },
    'Sep 2025': {
        'rate': 1400,
        'income': [],
        'fixed': [('Swiss', 150000, None), ('Olav Depto', 100000, None), ('Gym+Cal', 37000, None), ('Asisa', 50000, None), ('Santander Esp', 20000, None), ('Psico', 16000, None), ('Comida Olav', 600000, None)],
        'daily': [('08', 'Reme Patagonia', 43000, None), ('15', 'Havana', 50000, None)],
    },
    'Oct 2025': {
        'rate': 1450,
        'income': [],
        'fixed': [('Swiss', 165000, None), ('Olav Depto', 100000, None), ('Endomedicame', 50000, None), ('Gym', 37000, None), ('Asisa', 50000, None), ('Santander Esp', 20000, None), ('Psico', 16000, None), ('Comida Olav', 600000, None)],
        'daily': [],
    },
    'Nov 2025': {
        'rate': 1450,
        'income': [],
        'fixed': [('Comida y mas', 900000, None), ('Swiss', 165000, None), ('Olav Depto', 100000, None), ('Pilates', 43000, None), ('Gym', 34000, None), ('Asisa', 50000, None), ('Santander Esp', 20000, None), ('Psico', 16000, None), ('Tenis', 23000, None), ('Internet', 13748, None), ('Regalo Guido', 150000, None), ('Chapa', 250000, None), ('Sur Pesca', 350000, None)],
        'daily': [],
    },
    'Dec 2025': {
        'rate': 1500,
        'income': [('Basecoat', None, 2750)],
        'fixed': [('Comida y mas', 900000, None), ('Swiss Dic', 172000, None), ('Olav Depto', 150000, None), ('Gym', 34000, None), ('Psico Anab', 30000, None), ('Santander Esp', 20000, None), ('Tenis', 23000, None), ('Internet', 13748, None)],
        'daily': [],
    },
    'Jan 2026': {
        'rate': 1500,
        'income': [('Basecoat', 4500000, 3000)],
        'fixed': [('Comida y mas', 900000, 600), ('Swiss Dic', 172000, 114.67), ('Olav Depto', 150000, 100), ('Gym', 34000, 22.67), ('Psico Anab', 30000, 20), ('Santander Esp', 20000, 13.33), ('Otorrino', 60000, 40), ('Tenis', 23000, 15.33), ('Internet', 13748, 9.17), ('Alquiler Cancha Tenis', 24000, 16), ('Haras Jodita', 25000, 16.67), ('Racing x3', 30000, 20)],
        'daily': [('01', 'Shell', 59000, None), ('02', 'Chapita', 40000, None), ('03', 'Cosmos', 90000, None), ('04', 'Venti Cena', 220000, None), ('05', 'Casa', 225000, None)],
    },
    'Feb 2026': {
        'rate': 1500,
        'income': [('Basecoat', 4500000, 3000)],
        'fixed': [('Comida y mas', 900000, 600), ('Swiss Dic', 172000, 114.67), ('Olav Depto', 150000, 100), ('Gym', 34000, 22.67), ('Psico Anab', 30000, 20), ('Santander Esp', 20000, 13.33), ('Otorrino', 60000, 40), ('Tenis', 23000, 15.33), ('Internet', 13748, 9.17), ('Alquiler Cancha Tenis', 24000, 16), ('Haras Jodita', 25000, 16.67), ('Racing x3', 30000, 20)],
        'daily': [('01', 'Shell', 59000, None), ('02', 'Chapita', 40000, None), ('03', 'Cosmos', 90000, None), ('04', 'Venti Cena', 220000, None), ('05', 'Casa', 225000, None)],
    },
    'Mar 2026': {
        'rate': 1500,
        'income': [('Basecoat', 3750000, 2500)],
        'fixed': [('Swiss Dic', 194000, 129.33), ('Olav Depto', 150000, 100), ('Gym Ohana', 34000, 22.67), ('Psico Anab', 30000, 20), ('Monotributo', 50000, 33.33), ('Contadora', 32000, 21.33), ('Internet', 26000, 17.33), ('Bubalo Tragos', 46000, 30.67), ('Bubalo Entrada', 25000, 16.67), ('Pastillas', 39000, 26)],
        'daily': [('02', 'Compra ML Journal', 40000, None)],
    },
    'Apr 2026': {
        'rate': 1500,
        'income': [('Basecoat', 3750000, 2500)],
        'fixed': [('Swiss Dic', 194000, 129.33), ('Olav Depto', 150000, 100), ('Gym Ohana', 34000, 22.67), ('Psico Anab', 30000, 20), ('Monotributo', 50000, 33.33), ('Contadora', 32000, 21.33), ('Psic Cognitivo', 70000, 46.67), ('Internet', 26000, 17.33), ('Bubalo Tragos', 46000, 30.67), ('Bubalo Entrada', 25000, 16.67), ('Pastillas', 39000, 26), ('Gastos Abril', 2200000, 1466.67)],
        'daily': [('02', 'Compra ML Journal', 40000, None)],
    },
    'May 2026': {
        'rate': 1500,
        'income': [('Basecoat', 3750000, 2500)],
        'fixed': [('Swiss Dic', 240000, 160), ('Gym Ohana', 34000, 22.67), ('Monotributo', 50000, 33.33), ('Psic test', 250000, 166.67), ('Psic Cognitivo', 70000, 46.67), ('Internet', 26000, 17.33)],
        'daily': [('02', 'Compra ML Journal', 40000, None), ('19', 'PS5 Joystick', None, 95)],
    },
}

OLD_DATA = {
    'Oct 2023': {
        'rate': 960,
        'income': [('Lucas Food Truck 35k', 35000, 36.46), ('R Zapas Tincho Vans Bordo', 30000, 31.25), ('D Dani 530', 502080, 523), ('R Casaca River Negra', 15000, 15.63), ('V Vapos x3', 32200, 33.54), ('F1 A Hs82 y Tom saez', 127400, 132.71), ('M MODEL PEYA', 200000, 208.33), ('A Iphone Gonza CF 13Pro 98%', 100800, 105), ('A Ipad 9na', 90000, 93.75)],
        'fixed': [('Rentas', 6000, 6.06, 'Pagado'), ('Exp Arenales', 9000, 9.09, 'Pagado'), ('AFIP', 5500, 5.56, 'Pagado'), ('Boca', 2500, 2.53, 'Auto'), ('Smiles', 3500, 3.54, 'Auto'), ('Gali Prestamo', 15500, 15.66, 'Pendiente'), ('BBVA Tarjeta', 10000, 10.10, 'Pagado')],
        'daily': [('01', '1er pago Cabaña', 50000, 52.08), ('02', 'Veggie', 4000, 4.17), ('03', '2do Pago Cabaña', 118000, 122.92), ('04', 'Tarta', 5000, 5.21), ('05', 'Compras Depto', 18000, 18.75), ('06', 'Pava Electrica', 7000, 7.29), ('07', 'Tapa de baño', 11000, 11.46), ('08', 'Casaca Barza', 7000, 7.29), ('09', 'Polleria', 10000, 10.42), ('10', 'Kiosco', 5000, 5.21)],
    },
    'Nov 2023': {
        'rate': 960,
        'income': [('D Dani', 500000, 500), ('R Oakley', 140000, 145.83), ('F Feba', 120000, 125), ('R Remera boca', 5000, 5.21), ('R Zapatillas nike', 40000, 41.67)],
        'fixed': [('HSBC', 275000, 275, 'Pendiente'), ('Expensas + Estela', 67000, 67.68, None), ('Gali Prestamo', 15500, 15.66, 'Pagado')],
        'daily': [('01', 'Remera Real Madrid', 8500, 8.85), ('02', 'Non Parei 1kg', 7000, 7.29), ('03', 'Comida x2', 4000, 4.17), ('05', 'Pizza', 3500, 3.65), ('08', 'Sanguche', 2500, 2.60), ('12', 'Kiosco', 3000, 3.13), ('15', 'Almuerzo', 5000, 5.21)],
    },
    'Dec 2023': {
        'rate': 960,
        'income': [('D Dani', 500000, 500), ('A Airbnb', 50000, 50), ('A BETA 11 al 15', 140000, 140)],
        'fixed': [('HSBC', 275000, 275, None), ('Expensas + Estela', 67000, 67.68, None), ('Gali Prestamo', 15500, 15.66, None)],
        'daily': [('01', 'Didi cumple pen', 2500, 2.60), ('02', 'Sang Vacio', 3000, 3.12), ('03', 'Emp Crazy che', 3200, 3.33), ('05', 'Kiosco', 2000, 2.08), ('10', 'Comida x2', 8000, 8.33), ('15', 'Almuerzo', 5000, 5.21), ('20', 'Cerveza', 3000, 3.12)],
    },
    'Jan 2024': {
        'rate': 1150,
        'income': [('R Mochila Columbia', 70000, 70), ('R Remera Real Madrid', 25000, 21.74), ('R Hamburgo', 22000, 19.13)],
        'fixed': [('HSBC', 200000, 202.02, None), ('Expensas + Estela', 99000, 99, None), ('Gali Prestamo', 15500, 15.66, None)],
        'daily': [('01', 'Didi cumple pen', 2500, 2.17), ('02', 'Frutas', 8000, 6.96), ('04', 'Corchio', 6000, 5.22)],
    },
    'Feb 2024': {
        'rate': 1150,
        'income': [('2 catta Valen', 65000, 56.52), ('1 publi Feba', 20000, 17.39)],
        'fixed': [('HSBC', 200000, 202.02, None), ('Gali Prestamo', 15500, 15.66, None), ('Rentas', 6000, 6.06, None)],
        'daily': [('01', 'Chip E-sim', 2500, 2.17)],
    },
    'Mar 2024': {
        'rate': 1,
        'income': [('AirbnbGift', None, 160), ('Amazon Marcel', None, 332), ('Jattin', None, 26), ('Helpling Horas', None, 270)],
        'fixed': [('BBVA VISA', None, 3500, None), ('BBVA MCARD', None, 4700, None), ('Monotrib', None, 11000, None)],
        'daily': [('09', 'Bolt', None, 11), ('10', 'O2 Recarga', None, 15), ('11', 'Anmeldung', None, 40), ('12', 'DB Tickets', None, 28), ('15', 'Supermercado', None, 45)],
    },
    'Apr 2024': {
        'rate': 1,
        'income': [('Andalucia Propi', None, 110), ('Helpling Abril', None, 684)],
        'fixed': [('Monotrib', None, 11000, None), ('BBVA VISA', None, 3500, None), ('BBVA MCARD', None, 4700, None)],
        'daily': [],
    },
    'May 2024': {
        'rate': 1,
        'income': [('Hellpling Ramona', None, 25), ('Bici SP blanca', None, 145), ('Bici Sp Elops', None, 120), ('Amazon Helpling', None, 200)],
        'fixed': [('DAK', None, 300, None), ('Monotrib', None, 11000, None), ('Exp Arenales', None, 42000, None)],
        'daily': [],
    },
    'Jun 2024': {
        'rate': 1,
        'income': [],
        'fixed': [],
        'daily': [('02', 'Cash USD', None, 200), ('03', 'Payo USD', None, 380), ('04', 'Revolut', None, 108), ('10', 'Comida', None, 35), ('15', 'Transport', None, 25)],
    },
    'Jul 2024': {
        'rate': 1,
        'income': [],
        'fixed': [],
        'daily': [('08', 'Burrito', None, 10), ('09', 'Pizza', None, 8), ('10', 'Kebab', None, 12), ('15', 'Supermercado', None, 45), ('20', 'Transporte', None, 18)],
    },
    'Aug 2024': {
        'rate': 1,
        'income': [],
        'fixed': [],
        'daily': [('01', 'Comida', None, 30), ('05', 'Transporte', None, 20), ('10', 'Supermercado', None, 55), ('15', 'Restaurante', None, 35), ('20', 'Misc', None, 25)],
    },
    'Sep 2024': {
        'rate': 1,
        'income': [],
        'fixed': [],
        'daily': [('02', 'Calzoni', None, 8), ('03', 'Pizza', None, 6), ('05', 'Remeras HM', None, 20), ('10', 'Supermercado', None, 40), ('15', 'Transporte', None, 15)],
    },
    'Oct 2024': {
        'rate': 1,
        'income': [],
        'fixed': [],
        'daily': [('04', 'Yerba', None, 8), ('06', 'Nachos y hummus', None, 9), ('07', 'Pizza Mercadona', None, 6), ('15', 'Supermercado', None, 38), ('20', 'Misc', None, 20)],
    },
    'Nov 2024': {
        'rate': 1200,
        'income': [],
        'fixed': [],
        'daily': [('16', 'Psico', None, 15), ('18', 'SUBE', 5000, None), ('19', 'Didi a ver a los bull', 5590, None), ('20', 'Almuerzo', 3000, None), ('25', 'Comida', 4000, None)],
    },
    'Dec 2024': {
        'rate': 1200,
        'income': [],
        'fixed': [],
        'daily': [('11', 'Sanguche', 7500, None), ('12', 'Tognis FIFA', 15000, None), ('13', 'Didi FIFA', 10000, None), ('14', 'Comida', 5000, None), ('20', 'Kiosco', 3000, None)],
    },
    'Jan 2025': {
        'rate': 1100,
        'income': [],
        'fixed': [],
        'daily': [('02', 'Rapanui', 6000, None), ('03', 'Didi', 4800, None), ('04', 'Echeverria pasaje', 14000, None), ('05', 'Comida', 8000, None), ('10', 'Almuerzo', 5000, None)],
    },
    'Feb 2025': {
        'rate': 1100,
        'income': [],
        'fixed': [],
        'daily': [],
    },
}

# ─── SHEET ORDER ──────────────────────────────────────────────────────────────

MONTH_ORDER = [
    'Oct 2023', 'Nov 2023', 'Dec 2023',
    'Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024', 'May 2024',
    'Jun 2024', 'Jul 2024', 'Aug 2024', 'Sep 2024', 'Oct 2024',
    'Nov 2024', 'Dec 2024',
    'Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025',
    'Jun 2025', 'Jul 2025', 'Aug 2025', 'Sep 2025', 'Oct 2025',
    'Nov 2025', 'Dec 2025',
    'Jan 2026', 'Feb 2026', 'Mar 2026', 'Apr 2026', 'May 2026',
]

# ─── HELPER FUNCTIONS ─────────────────────────────────────────────────────────

def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_idx, width)"""
    for col, w in widths:
        ws.column_dimensions[col].width = w

def style_cell(cell, fill_color=None, font_color=C_WHITE, bold=False,
               align="left", size=11, border=False, num_format=None):
    if fill_color:
        cell.fill = fill(fill_color)
    cell.font = font(font_color, bold, size)
    cell.alignment = center() if align == "center" else left()
    if border:
        cell.border = thin_border()
    if num_format:
        cell.number_format = num_format

def write_title(ws, month_label, row=1):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=f"💸 {month_label.upper()} — GASTOS MENSUALES")
    style_cell(c, C_NAVY, C_WHITE, bold=True, align="center", size=14)
    ws.row_dimensions[row].height = 30

def write_rate_row(ws, rate, row=2):
    ws.cell(row=row, column=1, value="USD/ARS")
    ws.merge_cells(f"A{row}:A{row}")
    c_label = ws.cell(row=row, column=1)
    style_cell(c_label, C_DARKBLUE, C_WHITE, bold=True, align="center")

    c_val = ws.cell(row=row, column=2, value=rate)
    style_cell(c_val, C_DARKBLUE, C_YELLOW, bold=True, align="center")

    c_note = ws.cell(row=row, column=3, value="← update monthly")
    style_cell(c_note, C_DARKBLUE, "AAAAAA", align="left", size=9)

    for col in range(4, 8):
        c = ws.cell(row=row, column=col)
        c.fill = fill(C_DARKBLUE)
    ws.row_dimensions[row].height = 22

def write_section_header(ws, label, color, row):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=label)
    style_cell(c, color, C_WHITE, bold=True, align="center", size=12)
    ws.row_dimensions[row].height = 24

def write_subheader(ws, headers, row, bg_color=C_ACCENT):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        style_cell(c, bg_color, C_WHITE, bold=True, align="center", size=10, border=True)
    ws.row_dimensions[row].height = 20

def write_gap(ws, row):
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.fill = fill("2A2A4A")
    ws.row_dimensions[row].height = 8

def write_total_row(ws, label, ars_col_letter, usd_col_letter,
                    start_data_row, end_data_row, total_row):
    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws.cell(row=total_row, column=1, value=label)
    style_cell(c, C_YELLOW, "1A1A2E", bold=True, align="center")

    if ars_col_letter:
        ars_col_num = openpyxl.utils.column_index_from_string(ars_col_letter)
        c_ars = ws.cell(row=total_row, column=ars_col_num)
        c_ars.value = f"=SUM({ars_col_letter}{start_data_row}:{ars_col_letter}{end_data_row})"
        style_cell(c_ars, C_YELLOW, "1A1A2E", bold=True, align="center", num_format='#,##0.00')

    if usd_col_letter:
        usd_col_num = openpyxl.utils.column_index_from_string(usd_col_letter)
        c_usd = ws.cell(row=total_row, column=usd_col_num)
        c_usd.value = f"=SUM({usd_col_letter}{start_data_row}:{usd_col_letter}{end_data_row})"
        style_cell(c_usd, C_YELLOW, "1A1A2E", bold=True, align="center", num_format='#,##0.00')

    ws.row_dimensions[total_row].height = 22

def add_dropdowns(ws, dv_category, dv_payment, dv_status, daily_start, daily_end,
                  fixed_start, fixed_end):
    # Category (col F) and Payment (col G) for daily rows
    if daily_end >= daily_start:
        dv_category.add(f"F{daily_start}:F{daily_end}")
        dv_payment.add(f"G{daily_start}:G{daily_end}")
    # Status (col E) for fixed rows
    if fixed_end >= fixed_start:
        dv_status.add(f"E{fixed_start}:E{fixed_end}")

def make_dropdowns():
    dv_cat = DataValidation(
        type="list",
        formula1='"Food,Transport,Health,Utilities,Entertainment,Tax,Savings,Big Buy,Personal,Other"',
        allow_blank=True,
        showDropDown=False
    )
    dv_pay = DataValidation(
        type="list",
        formula1='"Cash ARS,USD Cash,Wise,Binance,MercadoPago,Tarjeta"',
        allow_blank=True,
        showDropDown=False
    )
    dv_status = DataValidation(
        type="list",
        formula1='"Pagado,Pendiente,Auto"',
        allow_blank=True,
        showDropDown=False
    )
    return dv_cat, dv_pay, dv_status

# ─── BUILD MONTH SHEET ────────────────────────────────────────────────────────

def build_month_sheet(wb, month_label, data):
    ws = wb.create_sheet(title=month_label)
    ws.sheet_view.showGridLines = False

    rate = data['rate']
    income_rows = data['income']
    fixed_rows = data['fixed']
    daily_rows = data['daily']

    # Column widths: A=6, B=11, C=32, D=14, E=14, F=18, G=16
    set_col_widths(ws, [
        ('A', 6), ('B', 11), ('C', 32), ('D', 14),
        ('E', 14), ('F', 18), ('G', 16)
    ])

    current_row = 1

    # ROW 1: Title
    write_title(ws, month_label, current_row)
    current_row += 1  # 2

    # ROW 2: Rate
    write_rate_row(ws, rate, current_row)
    current_row += 1  # 3

    # ROW 3: ▲ INGRESOS header
    write_section_header(ws, "▲ INGRESOS", C_GREEN, current_row)
    current_row += 1  # 4

    # ROW 4: Income sub-headers
    write_subheader(ws, ['#', 'Source', 'Description', 'ARS', 'USD'], current_row, C_GREEN)
    current_row += 1  # 5

    income_data_start = current_row
    # Income data rows
    for i, inc in enumerate(income_rows):
        desc, ars_val, usd_val = inc[0], inc[1], inc[2]
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW

        c_num = ws.cell(row=current_row, column=1, value=i + 1)
        style_cell(c_num, row_bg, C_DARKBLUE, align="center", border=True)

        c_src = ws.cell(row=current_row, column=2, value="")
        style_cell(c_src, row_bg, C_DARKBLUE, border=True)

        c_desc = ws.cell(row=current_row, column=3, value=desc)
        style_cell(c_desc, row_bg, C_DARKBLUE, border=True)

        # ARS column D
        c_ars = ws.cell(row=current_row, column=4)
        if ars_val is not None:
            c_ars.value = ars_val
        style_cell(c_ars, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        # USD column E
        c_usd = ws.cell(row=current_row, column=5)
        if usd_val is not None:
            c_usd.value = usd_val
        elif ars_val is not None:
            c_usd.value = f"=IF(D{current_row}>0,D{current_row}/B$2,\"\")"
        style_cell(c_usd, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        for col in range(6, 8):
            c = ws.cell(row=current_row, column=col)
            c.fill = fill(row_bg)

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    income_data_end = current_row - 1

    # TOTAL INGRESOS row
    income_total_row = current_row
    write_total_row(ws, "TOTAL INGRESOS", "D", "E",
                    income_data_start, income_data_end, income_total_row)
    for col in range(6, 8):
        c = ws.cell(row=income_total_row, column=col)
        c.fill = fill(C_YELLOW)
    current_row += 1

    # GAP
    write_gap(ws, current_row)
    current_row += 1

    # ▼ EGRESOS FIJOS header
    write_section_header(ws, "▼ EGRESOS FIJOS", C_RED, current_row)
    current_row += 1

    # Fixed sub-headers
    write_subheader(ws, ['#', 'Description', 'ARS', 'USD', 'Status', '', ''],
                    current_row, C_RED)
    current_row += 1

    fixed_data_start = current_row
    for i, fx in enumerate(fixed_rows):
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW

        # Handle both old (4-tuple with status) and new (3-tuple) formats
        if len(fx) == 4:
            desc, ars_val, usd_val, status = fx
        else:
            desc, ars_val, usd_val = fx
            status = None

        c_num = ws.cell(row=current_row, column=1, value=i + 1)
        style_cell(c_num, row_bg, C_DARKBLUE, align="center", border=True)

        c_desc = ws.cell(row=current_row, column=2, value=desc)
        # Merge B-C for description
        ws.merge_cells(f"B{current_row}:C{current_row}")
        style_cell(c_desc, row_bg, C_DARKBLUE, border=True)

        c_ars = ws.cell(row=current_row, column=4)
        if ars_val is not None:
            c_ars.value = ars_val
        style_cell(c_ars, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        c_usd = ws.cell(row=current_row, column=5)
        if usd_val is not None:
            c_usd.value = usd_val
        elif ars_val is not None:
            c_usd.value = f"=IF(D{current_row}>0,D{current_row}/B$2,\"\")"
        style_cell(c_usd, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        # Status col F (was E in spec, but col 6 = F here due to merged B:C)
        # Actually fixed section uses cols A,B-C(merged),D,E,F for status
        c_status = ws.cell(row=current_row, column=6, value=status or "")
        style_cell(c_status, row_bg, C_DARKBLUE, border=True)

        c_extra = ws.cell(row=current_row, column=7)
        c_extra.fill = fill(row_bg)

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    fixed_data_end = current_row - 1

    # TOTAL FIJOS
    fixed_total_row = current_row
    write_total_row(ws, "TOTAL FIJOS", "D", "E",
                    fixed_data_start, fixed_data_end, fixed_total_row)
    for col in range(6, 8):
        c = ws.cell(row=fixed_total_row, column=col)
        c.fill = fill(C_YELLOW)
    current_row += 1

    # GAP
    write_gap(ws, current_row)
    current_row += 1

    # 📋 GASTOS DEL MES header
    write_section_header(ws, "📋 GASTOS DEL MES", C_ACCENT, current_row)
    current_row += 1

    # Daily sub-headers: # | Date | Description | ARS | USD | Category | Payment
    write_subheader(ws, ['#', 'Date', 'Description', 'ARS', 'USD', 'Category', 'Payment'],
                    current_row, C_ACCENT)
    current_row += 1

    daily_data_start = current_row
    total_daily_rows = 25
    filled_rows = 0

    for i, day in enumerate(daily_rows):
        date_val, desc, ars_val, usd_val = day
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW
        filled_rows += 1

        c_num = ws.cell(row=current_row, column=1, value=i + 1)
        style_cell(c_num, row_bg, C_DARKBLUE, align="center", border=True)

        c_date = ws.cell(row=current_row, column=2, value=date_val)
        style_cell(c_date, row_bg, C_DARKBLUE, align="center", border=True)

        c_desc = ws.cell(row=current_row, column=3, value=desc)
        style_cell(c_desc, row_bg, C_DARKBLUE, border=True)

        c_ars = ws.cell(row=current_row, column=4)
        if ars_val is not None:
            c_ars.value = ars_val
        style_cell(c_ars, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        c_usd = ws.cell(row=current_row, column=5)
        if usd_val is not None:
            c_usd.value = usd_val
        elif ars_val is not None:
            c_usd.value = f"=IF(D{current_row}>0,D{current_row}/B$2,\"\")"
        style_cell(c_usd, row_bg, C_DARKBLUE, border=True, num_format='#,##0.00')

        # Category col F, Payment col G — blank with dropdown
        c_cat = ws.cell(row=current_row, column=6, value="")
        style_cell(c_cat, row_bg, C_DARKBLUE, border=True)

        c_pay = ws.cell(row=current_row, column=7, value="")
        style_cell(c_pay, row_bg, C_DARKBLUE, border=True)

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Blank rows to reach 25 total
    for i in range(filled_rows, total_daily_rows):
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW
        for col in range(1, 8):
            c = ws.cell(row=current_row, column=col)
            c.fill = fill(row_bg)
            c.border = thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    daily_data_end = current_row - 1

    # TOTAL VARIABLE
    variable_total_row = current_row
    write_total_row(ws, "TOTAL VARIABLE", "D", "E",
                    daily_data_start, daily_data_end, variable_total_row)
    for col in range(6, 8):
        c = ws.cell(row=variable_total_row, column=col)
        c.fill = fill(C_YELLOW)
    current_row += 1

    # GAP
    write_gap(ws, current_row)
    current_row += 1

    # 💰 BALANCE row
    balance_row = current_row
    ws.merge_cells(f"A{balance_row}:C{balance_row}")
    c_bl = ws.cell(row=balance_row, column=1, value="💰 BALANCE")
    style_cell(c_bl, C_NAVY, C_YELLOW, bold=True, align="center", size=12)

    # Balance ARS = Income ARS - Fixed ARS - Variable ARS
    c_bal_ars = ws.cell(row=balance_row, column=4)
    c_bal_ars.value = (f"=D{income_total_row}-D{fixed_total_row}-D{variable_total_row}")
    style_cell(c_bal_ars, C_NAVY, C_YELLOW, bold=True, align="center", num_format='#,##0.00')

    # Balance USD
    c_bal_usd = ws.cell(row=balance_row, column=5)
    c_bal_usd.value = (f"=E{income_total_row}-E{fixed_total_row}-E{variable_total_row}")
    style_cell(c_bal_usd, C_NAVY, C_YELLOW, bold=True, align="center", num_format='#,##0.00')

    for col in range(6, 8):
        c = ws.cell(row=balance_row, column=col)
        c.fill = fill(C_NAVY)
    ws.row_dimensions[balance_row].height = 26

    # Dropdowns
    dv_cat, dv_pay, dv_status = make_dropdowns()
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_pay)
    ws.add_data_validation(dv_status)
    add_dropdowns(ws, dv_cat, dv_pay, dv_status,
                  daily_data_start, daily_data_end,
                  fixed_data_start, fixed_data_end)

    # Freeze panes at row 4 (rows 1-3 frozen)
    ws.freeze_panes = "A4"

    return {
        'income_total_row': income_total_row,
        'fixed_total_row': fixed_total_row,
        'variable_total_row': variable_total_row,
        'balance_row': balance_row,
    }

# ─── BIG BUYS TAB ─────────────────────────────────────────────────────────────

BIG_BUYS = [
    ('Oct 2023', '01/10', '1er Pago Cabaña', None, 50000, ''),
    ('Oct 2023', '03/10', '2do Pago Cabaña', None, 118000, ''),
    ('Oct 2023', '??/10', 'Iphone Gonza CF 13Pro', 105, None, ''),
    ('Nov 2023', '??/11', 'D Dani (deuda)', 500, None, ''),
    ('Nov 2023', '??/11', 'HSBC pago', 275, None, ''),
    ('Mar 2025', '??/03', 'MSI Cyborg laptop', 810, None, ''),
    ('Mar 2025', '??/03', 'Sillon', None, 230000, ''),
    ('Mar 2025', '??/03', 'Monitor', None, 250000, ''),
    ('Nov 2025', '??/11', 'Sur Pesca', None, 350000, ''),
    ('Nov 2025', '??/11', 'Chapa', None, 250000, ''),
    ('Nov 2025', '??/11', 'Regalo Guido', None, 150000, ''),
    ('Apr 2026', '??/04', 'Gastos Abril (varios)', 1466.67, None, 'Big month'),
    ('May 2026', '19/05', 'PS5 Joystick (DualSense)', 95, None, ''),
]

def build_big_buys_sheet(wb):
    ws = wb.create_sheet(title="🛍 Big Buys")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [('A', 12), ('B', 10), ('C', 35), ('D', 12), ('E', 14), ('F', 20)])

    # Title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="🛍 ALL-TIME BIG PURCHASES")
    style_cell(c, C_NAVY, C_WHITE, bold=True, align="center", size=14)
    ws.row_dimensions[1].height = 30

    # Sub-headers
    headers = ['Month', 'Date', 'Item', 'USD', 'ARS', 'Notes']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        style_cell(c, C_ACCENT, C_WHITE, bold=True, align="center", border=True)
    ws.row_dimensions[2].height = 20

    data_start = 3
    for i, (month, date, item, usd, ars, notes) in enumerate(BIG_BUYS):
        row = i + data_start
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW

        ws.cell(row=row, column=1, value=month).fill = fill(row_bg)
        ws.cell(row=row, column=2, value=date).fill = fill(row_bg)
        ws.cell(row=row, column=3, value=item).fill = fill(row_bg)

        c_usd = ws.cell(row=row, column=4, value=usd)
        c_usd.fill = fill(row_bg)
        if usd:
            c_usd.number_format = '#,##0.00'

        c_ars = ws.cell(row=row, column=5, value=ars)
        c_ars.fill = fill(row_bg)
        if ars:
            c_ars.number_format = '#,##0.00'

        ws.cell(row=row, column=6, value=notes).fill = fill(row_bg)

        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.font = Font(color=C_DARKBLUE, size=10)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    total_row = data_start + len(BIG_BUYS)
    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    style_cell(c, C_YELLOW, C_DARKBLUE, bold=True, align="center")

    c_usd_total = ws.cell(row=total_row, column=4)
    c_usd_total.value = f"=SUM(D{data_start}:D{total_row-1})"
    style_cell(c_usd_total, C_YELLOW, C_DARKBLUE, bold=True, align="center", num_format='#,##0.00')

    c_ars_total = ws.cell(row=total_row, column=5)
    c_ars_total.value = f"=SUM(E{data_start}:E{total_row-1})"
    style_cell(c_ars_total, C_YELLOW, C_DARKBLUE, bold=True, align="center", num_format='#,##0.00')

    ws.cell(row=total_row, column=6).fill = fill(C_YELLOW)
    ws.row_dimensions[total_row].height = 22

    ws.freeze_panes = "A3"

# ─── CASH ASSETS TAB ──────────────────────────────────────────────────────────

CASH_ASSETS = [
    ('Wise (USD)', 14800, ''),
    ('USD Cash (Olav)', 323, ''),
    ('EUR Cash (Olav)', 550 * 1.08, '× 1.08 USD/EUR'),  # 594 USD
    ('Binance USDT', 0, ''),
    ('Nexo', 0, ''),
    ('MercadoPago', None, ''),
    ('Cash ARS', 60000 / 1500, '÷ 1500 ARS/USD'),  # 40 USD
    ('Terreno (land)', 30000, 'USD'),
    ('US Bank', 45, ''),
    ('Level', 300, ''),
]

def build_cash_assets_sheet(wb):
    ws = wb.create_sheet(title="💰 Cash Assets")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [('A', 22), ('B', 14), ('C', 25)])

    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="💰 CASH ASSETS — MAY 2026")
    style_cell(c, C_NAVY, C_WHITE, bold=True, align="center", size=14)
    ws.row_dimensions[1].height = 30

    headers = ['Asset', 'USD Value', 'Notes']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        style_cell(c, C_ACCENT, C_WHITE, bold=True, align="center", border=True)
    ws.row_dimensions[2].height = 20

    data_start = 3
    for i, (asset, val, notes) in enumerate(CASH_ASSETS):
        row = i + data_start
        row_bg = C_LIGHTGRAY if i % 2 == 0 else C_ALTROW

        c_asset = ws.cell(row=row, column=1, value=asset)
        c_asset.fill = fill(row_bg)
        c_asset.font = Font(color=C_DARKBLUE, bold=False, size=11)
        c_asset.alignment = left()
        c_asset.border = thin_border()

        c_val = ws.cell(row=row, column=2, value=val)
        c_val.fill = fill(row_bg)
        c_val.font = Font(color=C_DARKBLUE, size=11)
        c_val.alignment = center()
        c_val.border = thin_border()
        if val is not None:
            c_val.number_format = '#,##0.00'

        c_notes = ws.cell(row=row, column=3, value=notes)
        c_notes.fill = fill(row_bg)
        c_notes.font = Font(color="888888", size=10)
        c_notes.alignment = left()
        c_notes.border = thin_border()

        ws.row_dimensions[row].height = 20

    total_row = data_start + len(CASH_ASSETS)
    ws.merge_cells(f"A{total_row}:A{total_row}")
    c_lbl = ws.cell(row=total_row, column=1, value="TOTAL ≈ USD")
    style_cell(c_lbl, C_YELLOW, C_DARKBLUE, bold=True, align="center")

    c_total = ws.cell(row=total_row, column=2)
    c_total.value = f"=SUM(B{data_start}:B{total_row-1})"
    style_cell(c_total, C_YELLOW, C_DARKBLUE, bold=True, align="center", num_format='#,##0.00')

    c_empty = ws.cell(row=total_row, column=3)
    c_empty.fill = fill(C_YELLOW)
    ws.row_dimensions[total_row].height = 22

    ws.freeze_panes = "A3"

# ─── ANUAL TAB ────────────────────────────────────────────────────────────────

# Year → months in that year (label → months list)
YEAR_MONTHS = {
    '2023': ['Oct 2023', 'Nov 2023', 'Dec 2023'],
    '2024': ['Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024', 'May 2024',
             'Jun 2024', 'Jul 2024', 'Aug 2024', 'Sep 2024', 'Oct 2024',
             'Nov 2024', 'Dec 2024'],
    '2025': ['Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025',
             'Jun 2025', 'Jul 2025', 'Aug 2025', 'Sep 2025', 'Oct 2025',
             'Nov 2025', 'Dec 2025'],
    '2026': ['Jan 2026', 'Feb 2026', 'Mar 2026', 'Apr 2026', 'May 2026'],
}

def build_anual_sheet(wb, month_row_map):
    ws = wb.create_sheet(title="📊 Anual")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [('A', 10), ('B', 22), ('C', 22), ('D', 22), ('E', 18)])

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="📊 RESUMEN ANUAL")
    style_cell(c, C_NAVY, C_WHITE, bold=True, align="center", size=14)
    ws.row_dimensions[1].height = 30

    headers = ['Año', 'Total Ingresos (USD)', 'Total Egresos (USD)', 'Balance (USD)', 'Meses']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        style_cell(c, C_ACCENT, C_WHITE, bold=True, align="center", border=True)
    ws.row_dimensions[2].height = 20

    current_row = 3
    year_rows = []

    for year, months in YEAR_MONTHS.items():
        # Build cross-sheet SUM formulas
        income_parts = []
        expense_parts = []

        for month in months:
            if month in month_row_map:
                rows = month_row_map[month]
                # Use the sheet name with proper quoting
                safe_name = month.replace("'", "''")
                income_parts.append(f"'{safe_name}'!E{rows['income_total_row']}")
                expense_parts.append(
                    f"'{safe_name}'!E{rows['fixed_total_row']}+'{safe_name}'!E{rows['variable_total_row']}"
                )

        row_bg = C_LIGHTGRAY if len(year_rows) % 2 == 0 else C_ALTROW

        c_year = ws.cell(row=current_row, column=1, value=int(year))
        style_cell(c_year, row_bg, C_DARKBLUE, bold=True, align="center", border=True)

        c_inc = ws.cell(row=current_row, column=2)
        if income_parts:
            c_inc.value = "=" + "+".join(income_parts)
        else:
            c_inc.value = 0
        style_cell(c_inc, row_bg, C_DARKBLUE, align="center", border=True, num_format='#,##0.00')

        c_exp = ws.cell(row=current_row, column=3)
        if expense_parts:
            c_exp.value = "=" + "+".join(expense_parts)
        else:
            c_exp.value = 0
        style_cell(c_exp, row_bg, C_DARKBLUE, align="center", border=True, num_format='#,##0.00')

        c_bal = ws.cell(row=current_row, column=4)
        c_bal.value = f"=B{current_row}-C{current_row}"
        style_cell(c_bal, row_bg, C_DARKBLUE, align="center", border=True, num_format='#,##0.00')

        c_months = ws.cell(row=current_row, column=5, value=", ".join(months))
        style_cell(c_months, row_bg, "888888", align="left", size=9, border=True)

        ws.row_dimensions[current_row].height = 20
        year_rows.append(current_row)
        current_row += 1

    # Grand total
    ws.merge_cells(f"A{current_row}:A{current_row}")
    c_gt = ws.cell(row=current_row, column=1, value="TOTAL")
    style_cell(c_gt, C_YELLOW, C_DARKBLUE, bold=True, align="center")

    for col in range(2, 5):
        col_letter = get_column_letter(col)
        c = ws.cell(row=current_row, column=col)
        c.value = f"=SUM({col_letter}3:{col_letter}{current_row-1})"
        style_cell(c, C_YELLOW, C_DARKBLUE, bold=True, align="center", num_format='#,##0.00')

    c_empty = ws.cell(row=current_row, column=5)
    c_empty.fill = fill(C_YELLOW)
    ws.row_dimensions[current_row].height = 24

    ws.freeze_panes = "A3"

# ─── MERGE NEW EXPENSES ───────────────────────────────────────────────────────

NEW_EXPENSES_PATH = "/Users/guidocarminatti/Projects/playground/.tmp/new_expenses.json"

def load_new_expenses():
    """Load pending expenses from the web form import file."""
    from pathlib import Path
    p = Path(NEW_EXPENSES_PATH)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text())
    except Exception:
        return {}

def merge_month_data(base, new_entries):
    """Merge new_entries dict into a copy of base month data."""
    data = copy.deepcopy(base)
    for section in ('income', 'fixed', 'daily'):
        for entry in new_entries.get(section, []):
            if section == 'daily':
                # entry: {"day": "25", "desc": "Pizza", "ars": 15000, "usd": null}
                data['daily'].append((
                    entry.get('day', ''),
                    entry.get('desc', ''),
                    entry.get('ars') or None,
                    entry.get('usd') or None,
                ))
            elif section == 'fixed':
                # entry: {"desc": "Gym", "ars": 34000, "usd": null, "status": "Pagado"}
                row = (entry.get('desc', ''), entry.get('ars') or None, entry.get('usd') or None)
                if entry.get('status'):
                    row = row + (entry['status'],)
                data['fixed'].append(row)
            elif section == 'income':
                # entry: {"desc": "Basecoat", "ars": null, "usd": 2500}
                data['income'].append((
                    entry.get('desc', ''),
                    entry.get('ars') or None,
                    entry.get('usd') or None,
                ))
    return data

# ─── MAIN BUILD ───────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    new_expenses = load_new_expenses()
    if new_expenses:
        months_with_new = list(new_expenses.keys())
        print(f"  ℹ Merging new entries for: {', '.join(months_with_new)}")

    month_row_map = {}

    print("Building month sheets...")
    for month_label in MONTH_ORDER:
        # Determine data source
        if month_label in OLD_DATA:
            data = OLD_DATA[month_label]
        elif month_label in OLA_DATA:
            data = OLA_DATA[month_label]
        else:
            # Empty sheet
            data = {'rate': 1500, 'income': [], 'fixed': [], 'daily': []}

        if month_label in new_expenses:
            data = merge_month_data(data, new_expenses[month_label])

        rows = build_month_sheet(wb, month_label, data)
        month_row_map[month_label] = rows

        # Hide pre-2026 tabs (data intact, just not visible)
        year = month_label.split()[-1]
        if year in ('2023', '2024', '2025'):
            wb[month_label].sheet_state = 'hidden'

        print(f"  ✓ {month_label}")

    print("Building Big Buys sheet...")
    build_big_buys_sheet(wb)

    print("Building Cash Assets sheet...")
    build_cash_assets_sheet(wb)

    print("Building Anual sheet...")
    build_anual_sheet(wb, month_row_map)

    wb.save(OUTPUT_PATH)
    print(f"\n✅ Done! Saved to: {OUTPUT_PATH}")
    print(f"   Total sheets: {len(wb.sheetnames)}")
    print(f"   Month tabs: {len(MONTH_ORDER)}")
    print(f"   Other tabs: 3 (Big Buys, Cash Assets, Anual)")
    print(f"\nSheet list: {wb.sheetnames}")

if __name__ == "__main__":
    main()
